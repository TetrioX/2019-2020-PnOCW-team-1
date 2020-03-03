from os import walk
from openpyxl import Workbook, load_workbook, cell as xlcell, worksheet
import time
from colorsys import rgb_to_hsv
from colorsys import rgb_to_hls

def hexToRGB(h): # tuple (R,G,B)
    h = h.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


global ws
condition = [None, None]
result = []
keys = []
wb = load_workbook(filename = './result.xlsx')
print("Hoppaaaaaa")
ws = wb.active

image = 2


def within_range(bounds: tuple, cell: xlcell) -> bool:
    column_start, row_start, column_end, row_end = bounds
    row = cell.row
    if row >= row_start and row <= row_end:
        column = cell.column
        if column >= column_start and column <= column_end:
            return True
    return False

def get_value_merged(sheet: worksheet, cell: xlcell) -> any:
    for merged in sheet.merged_cells:
        if within_range(merged.bounds, cell):
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value

def select_brightness(value):
    condition[0] = "B"
    condition[1] = value

def select_distance(value):
    condition[0] = "E"
    condition[1] = value

def select_nb_colors(value):
    condition[0] = "F"
    condition[1] = value

def select_reflection(value):
    condition[0] = "H"
    condition[1] = value

def select_incidence(value):
    condition[0] = "G"
    condition[1] = value


def read_values():
    i = 2
    cond = condition[0]
    val = condition[1]
    print("cond=", cond)
    print('val=', val)
    while ws["L" + str(i)].value is not None:
        if get_value_merged(ws, ws[cond + str(i)]) == str(val):
            result.append(ws["L" + str(i)].value)
            keys.append(get_value_merged(ws, ws["K" + str(i)]))
        i += 1


select_brightness(50)
print("condition = ", condition)
read_values()
print(result)


global ws2
wb2 = load_workbook(filename = './statistics_HSL.xlsx')
ws2 = wb2.active

"""
ws2.cell(row=1, column=1).value = "R_actual"
ws2.cell(row=1, column=2).value = "G_actual"
ws2.cell(row=1, column=3).value = "B_actual"
ws2.cell(row=1, column=4).value = "R_measured"
ws2.cell(row=1, column=5).value = "G_measured"
ws2.cell(row=1, column=6).value = "B_measured"

"""

def add_values(actual, measured, n):
    """
    R, G en B zijn de echte waarden.
    Rm, Gm en Bm zijn de gemeten waarden.
    """
    actual = hexToRGB(actual)
    measured = hexToRGB(measured)

    actual = rgb_to_hls(actual[0], actual[1], actual[2])
    measured = rgb_to_hls(measured[0], measured[1], measured[2])

    R = actual[0]
    G = actual[1]
    B = abs(actual[2])
    Rm = measured[0]
    Gm = measured[1]
    Bm = abs(measured[2])


    ws2.cell(row=n, column=1).value = R
    ws2.cell(row=n, column=2).value = G
    ws2.cell(row=n, column=3).value = B
    ws2.cell(row=n, column=4).value = Rm
    ws2.cell(row=n, column=5).value = Gm
    ws2.cell(row=n, column=6).value = Bm


def all_values(actual_values, measured_values):
    i = 0
    while i < min(len(actual_values), len(measured_values)):
        add_values(actual_values[i], measured_values[i], i+2)
        i += 1



inputlist = ['#cb6e17', '#d06400', '#ce5300', '#de6a00', '#d98400', '#efbe04', '#b47b00', '#d9a909', '#b48200', '#e8c119', '#dcba0d', '#bd9402', '#ae8605', '#e7ce1a', '#e2c723', '#c1990a', '#b37a00', '#e4b107', '#f3b608', '#ce9100', '#3b61ef', '#4264f3', '#1c3bd2', '#303b9d', '#1a35c5', '#1131ca', '#2946dd', '#626de6', '#556eee', '#515ddb', '#2041d6', '#3f65ed', '#6e80d9', '#2d49cc', '#363fbe', '#6468de', '#3c5ae7', '#1d43df', '#3f5ce7', '#2d43c5', '#d74dd4', '#ee6ce9', '#eb6ae9', '#d959e0', '#d758dd', '#cf48c6', '#eb68db', '#f079db', '#ea6ad3', '#b53681', '#a63273', '#bf3e81', '#da5da7', '#d653aa', '#ac3988', '#a72f8e', '#9b3293', '#8147a6', '#a74ecd', '#b64bdd', '#3c56f0', '#9254ee', '#b54bda', '#d76def', '#d36ff1', '#ae56e9', '#465dd7', '#3e4dea', '#3b4de8', '#8142de', '#1932c8', '#8450ea', '#8120b6', '#e177ef', '#c25ce9', '#aa43c2', '#6c30c4', '#3f4de6', '#3f4de6', '#5063d6', '#9e9fd9', '#7779d7', '#826ad6', '#6585d8', '#b9dbeb', '#6ca2d1', '#8d9275', '#c7d1c4', '#d6f9ff', '#93b9c5', '#5e73cf', '#7ba6b5', '#8a80e4', '#7782d2', '#bd8dd5', '#eacfcb', '#947ecc', '#6b90ca', '#7672db', '#548acb', '#0e7364', '#5dd3a3', '#208075', '#0a6e68', '#136062', '#47b49e', '#005d49', '#186e62', '#146d5d', '#005e42', '#066b55', '#4a826b', '#094239', '#4f9480', '#227b89', '#105548', '#168275', '#1a6555', '#0d594d', '#207d7e', '#e89c5a', '#dc700a', '#ff8703', '#eeab02', '#fd8e04', '#f1b700', '#eb7c00', '#cd7d02', '#cd4e00', '#d64200', '#f86802', '#fd7303', '#e57c33', '#db9353', '#ac4900', '#a15a10', '#f1ad00', '#fe9700', '#ebb203', '#ffb00a', '#7d97cd', '#a0b1c4', '#77598e', '#724e78', '#538994', '#478a7b', '#9edad2', '#a0c299', '#95ad73', '#637948', '#3d6647', '#3e7086', '#454575', '#4e4685', '#595ca0', '#8ab3d4', '#90d3ce', '#52a3a0', '#89b1cd', '#809ecf', '#f27604', '#fe901a', '#e36b48', '#f27a02', '#f78000', '#ce6516', '#d04e0d', '#ec887c', '#c5535a', '#943036', '#c6585e', '#d56c75', '#cc5760', '#f7892f', '#fb8900', '#d65200', '#ec7635', '#f47f00', '#e46821', '#b14649', '#367f36', '#1a874f', '#1a874f', '#1a874f', '#1a874f', '#2aa450', '#26a556', '#128738', '#0f7c4e', '#2f936f', '#429378', '#456b5f', '#234e3f', '#123d30', '#124a37', '#5f7949', '#2e5e1c', '#1d775a', '#2fa669', '#2ba84c', '#ee810e', '#ff8b03', '#de5500', '#f57b00', '#f46a00', '#e35203', '#e45000', '#ff7b01', '#fa7800', '#e0600d', '#d0500f', '#d54602', '#b23005', '#c53401', '#ff8c00', '#f86c07', '#e54b04', '#fb7d00', '#ff8600', '#ff8d01', '#c53f99', '#f786d6', '#da5db9', '#cf50ab', '#d253a2', '#ef80ca', '#e66bc1', '#d354a9', '#db5cb2', '#d14ea9', '#dc5cad', '#e260bf', '#db67ba', '#da5dac', '#e060b9', '#d866a2', '#cf48a4', '#d85ab3', '#e869be', '#c745a4', '#dd956d', '#e47249', '#e287b8', '#e08fc0', '#c3699e', '#cd6e3c', '#d6643f', '#cb5942', '#c76191', '#cd7bb0', '#9b4b76', '#b85f62', '#f6976c', '#e47e5b', '#b44a73', '#b35a8b', '#cb79ad', '#954973', '#c96773', '#ed865f', '#dbd120', '#dcd129', '#99c843', '#50982e', '#c3e037', '#53a02d', '#bdb111', '#d7c425', '#ceb928', '#bbcb26', '#b2db2c', '#bfdb34', '#bfdb34', '#c5c710', '#e3d626', '#e1e221', '#bad726', '#9cb623', '#a89411', '#a38f1a', '#cfc623', '#e7da02', '#9e9407', '#c7c202', '#cdc702', '#d1c807', '#bbb701', '#d4cd01', '#eadf01', '#d6d202', '#c6c101', '#d5cd08', '#837608', '#867d00', '#d8d105', '#e5da02', '#c1bd00', '#ece502', '#bab502', '#897e04', '#179e30', '#01a125', '#00991c', '#02aa22', '#03b026', '#036d14', '#0a5118', '#0d7c1a', '#06871c', '#083d0b', '#006d11', '#066816', '#099523', '#049a22', '#17741f', '#05751d', '#125714', '#006811', '#00ac22', '#02a924', '#059e31', '#00a429', '#079a2c', '#0b9145', '#16944c', '#257a46', '#0b7b3b', '#00711e', '#08651a', '#089625', '#03991d', '#0a8b1e', '#176e18', '#1d5f18', '#017522', '#045026', '#144a2d', '#1f6e3d', '#018629', '#05a22f', '#4ac54d', '#006f02', '#2c810d', '#006b00', '#138513', '#138513', '#27ac3e', '#30a445', '#006a02', '#196400', '#598306', '#477f18', '#0a7e13', '#007f15', '#221300', '#3d4f00', '#617514', '#096f03', '#3c6d0b', '#006e06']
inputcolor = ['#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#00aaff', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff0000', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff00aa', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff0055', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ff5500', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#ffff00', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#0055ff', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#ffaa00', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#5500ff', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#aaff00', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#0000ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#ff00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#aa00ff', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffaa', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ffff', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#00ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#55ff00', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55', '#00ff55']




all_values(result, keys)



wb2.save('statistics_HSL.xlsx')


