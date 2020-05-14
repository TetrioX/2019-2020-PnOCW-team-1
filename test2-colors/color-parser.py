# Dependency: openpyxl, pillow, opencv-python

from os import walk, path
import cv2
from openpyxl import Workbook, load_workbook, cell, worksheet
import time
import math


### Config
# number of picks for each color
global nbp
nbp = 20

global ws
if (path.isfile("./result.xlsx")) and input("An existing result.xlsx file has been found. \
Do you want to use this file? (Y/n)\n").lower() != "n":
    print("==> reading result.xlsx (this can take several minutes)...", flush=True)
    wb = load_workbook(filename = './result.xlsx', data_only=True)
else:
    wb = Workbook()
ws = wb.active

# header
ws["A1"] = "group"
ws["B1"] = "brigthness"
ws["C1"] = "pictureDevice"
ws["D1"] = "ratioScreenPicture"
ws["E1"] = "distance"
ws["F1"] = "nbOfColros"
ws["G1"] = "incidence"
ws["H1"] = "reflection"
ws["I1"] = "screenDevice"
ws["J1"] = "imageID"
ws["K1"] = "origHexCol"
ws["L1"] = "picHexCol"

image = 2 # current row

def mergeRows(col, row, nb):
    if nb != 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row+nb-1, end_column=col)

def pickColor(file, pictureName, picNum, row):
    global pressedkey, nbp, cnbp
    image = cv2.imread(file)
    name = file.split("/")[-1]
    cnbp = 0
    sratio = min(100, int(pictureName[2]))
    nbOfColors = int(pictureName[4])
    if (sratio < 100 and nbOfColors > 2) or (sratio < 50 and nbOfColors >= 1):
        cv2.namedWindow(name, cv2.WND_PROP_FULLSCREEN)
        cv2.setWindowProperty(name,cv2.WND_PROP_FULLSCREEN,cv2.WINDOW_FULLSCREEN)
        cv2.imshow(name, image)
        cv2.setMouseCallback(name, getColor, (image, row))
        while (cnbp < nbp):
            pressedkey = False
            while pressedkey == False:
                cv2.waitKey(10)
                time.sleep(0.010)
            cnbp += 1
        cv2.destroyAllWindows()
    else:
        eratio = (100 - sratio)/150 + 0.05
        height, width = image.shape[:2]
        nbOfRows = math.ceil(math.sqrt(nbp))
        nbOfCols = math.ceil(nbp/nbOfRows)
        eheight = eratio * height
        ewidth = eratio * width
        hsize = (height - eheight*2)/(nbOfRows + 5)
        wsize = (width - ewidth*2)/(nbOfCols + 5)/nbOfColors
        wcsize = (width - ewidth*2)/nbOfColors
        for w in range(3, 3 + nbOfCols):
            for h  in range(3, 3 + nbOfRows):
                if cnbp >= nbp:
                    break
                x = int(ewidth + wcsize*picNum + wsize*w)
                y = int(eheight + hsize*h)
                row[-1] = rgbToHex(image[y, x][2], image[y, x][1], image[y, x][0])
                ws.append(row)
                cnbp += 1


def getColor(event,x,y,flags,param):
    global pressedkey, ws, cnbp
    if event == cv2.EVENT_LBUTTONDOWN:
        param[1][-1] = rgbToHex(param[0][y, x][2], param[0][y, x][1], param[0][y, x][0])
        ws.append(param[1])
        pressedkey = True
    if event == cv2.EVENT_RBUTTONDOWN:
        if input("Do you want to save the file? (Y/n)\n") != "n":
            wb.save('result.xlsx')
        if input("Do you want to exit?(y/N)?") == "y":
            quit()



def rgbToHex(r, g, b):
    return "#{0:02x}{1:02x}{2:02x}".format(int(r), int(g), int(b))


def get_hash(groupName, fileName):
    return groupName + "/" + fileName

# look for old colors
print("==> looking for old picks...", flush=True)
pickedList = set()
while ws["L"+str(image)].value is not None:
    group = ws["A"+str(image)].value
    file = ws["J"+str(image)].value
    if group is not None and file is not None:
        pickedList.add(get_hash(group, file))
    image += 1
ws.delete_rows(image)
oldMax = image
# pick new colors
print("==> picking new colors...", flush=True)
for group in walk("./"):
    if (group[0] == "./"):
        continue
    for pic in group[2]:
        groupName = group[0].split("/")[-1]
        fileName = pic.split(".")[0]
        pictureName = fileName.split("_")
        if len(pictureName) < 12 or (len(pictureName) - 9)%3 != 0 or (len(pictureName) - 9)//3 < int(pictureName[4]):
            print(group[0] + "/" + pic + " is poorly formatted")
            continue
        hash = get_hash(groupName, fileName)
        if hash in pickedList:
            print(group[0] + "/" + pic + " has already been picked")
            pickedList.remove(hash)
            continue
        nbc = int(pictureName[4])
        blocksize = nbp*nbc
        row = [groupName, pictureName[0], pictureName[1], pictureName[2], pictureName[3],\
        pictureName[4], pictureName[-4], pictureName[-3], pictureName[-2], fileName, None, None]
        for i in range(nbc):
            row[-2] = rgbToHex(pictureName[5 + i * 3], pictureName[6 + i * 3], pictureName[7 + i * 3])
            pickColor(group[0]+'/'+pic, pictureName, i, row)
        #    mergeRows(11, image + i*nbp, nbp)
        # for i in range(1, 11):
        #    mergeRows(i, image, nbp*nbc)
        print(group[0] + "/" + pic + " has been added")
        image += blocksize

if len(pickedList) != 0:
    print("The following files were already in excel but were not find in the actual data set. They might have been renamed/removed:\n ")
    for hash in pickedList:
        print(hash)
    if input("Do you want to remove these files? (y/N)\n") == "y":
        print("==> deleting old entries...", flush=True)
        image = 2
        while len(pickedList) > 0:
            group = ws["A"+str(image)].value
            file = ws["J"+str(image)].value
            if group is not None and file is not None and get_hash(groupName, fileName) in pickedList:
                ws.delete_rows(image)
            else:
                image += 1
            if image > oldMax:
                if input("ERROR: Couldn't remove all old entries. This is most likely a bug. Do you still want to save?(Y/n)\n") == "n":
                    quit()
                break


print("==> saving excel file")
wb.save('result.xlsx')
