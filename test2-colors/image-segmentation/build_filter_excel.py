# Dependency: openpyxl, pillow, opencv-python

from os import walk, path
import os
import cv2
from openpyxl import Workbook, load_workbook, cell, worksheet
import time
import math
import image_processing as ip
import tkinter as tk
from tkinter import filedialog
import sys
from PIL import ImageTk, Image
import colorsys







global nbp, ws_names, wb_names
nbp = 20
result = []
coord_dict = dict()

# SET UP EXCEL FILES

wb_names = ['NoAddedNoise', 'S&P', 'GaussianNoise', 'Speckle', 'Poisson']
ws_names = ['NoFilter', 'MedianFilter', 'GaussianFilter', 'MeanFilter', 'BilateralFilter']
pr = './'
exc = '.xlsx'
"""
for wbn in wb_names:
    filename = pr + wbn + exc
    wb = Workbook()
    for wsn in ws_names:
        sh = wb.create_sheet(wsn)
    wb.save(filename)
"""

# HANDLE IMAGES

dirNames = ['NoFilter', 'MedianFilter', 'GaussianFilter', 'MeanFilter', 'BilateralFilter']
noiseTypes = ['NoAddedNoise', 'S&P', 'GaussianNoise', 'Speckle', 'Poisson']

'''
tempp = './images'
for di in os.listdir(tempp):
    for d in dirNames:
        p = tempp + '/' + d
        for n in noiseTypes:
            pp = p + '/' + n
            if not os.path.exists(pp):
                os.mkdir(pp)
                print("Directory " , pp,  " Created ")
            else:
                print("Directory " , pp,  " already exists")
'''


img_dir = 'raw_input_img'
out_dir_noise = '.'
out_dir_filter = 'images'
ext = '.jpg'


def add_noise(path_in, out_dir, noise, i=0):
    rgb = input_img_getRGB(path_in)
    if noise == 'NoAddedNoise':
        path_out = out_dir_noise + '/input_img/' + noise + '/' + rgb + str(i) + ext
        im = cv2.imread(path_in)
        cv2.imwrite(path_out, im)
    elif noise == 'S&P':
        path_out = out_dir_noise + '/input_img/' + noise + '/' + rgb + str(i) + ext
        ip.salt_and_pepper(path_in, path_out)
    elif noise == 'GaussianNoise':
        path_out = out_dir_noise + '/input_img/' + noise + '/' + rgb + str(i) + ext
        ip.gauss_noise(path_in, path_out)
    elif noise == 'Speckle':
        path_out = out_dir_noise + '/input_img/' + noise + '/' + rgb + str(i) + ext
        ip.speckle(path_in, path_out)
    elif noise == 'Poisson':
        path_out = out_dir_noise + '/input_img/' + noise + '/' + rgb + str(i) + ext
        ip.poisson(path_in, path_out)
    else:
        print("NO VALID NOISE")

def apply_filter(path_in, out_dir, filter, noise, filename,i=0):
    if filter == 'GaussianFilter':
        path_out = out_dir_filter + '/' + filter + '/' + noise + '/' + filename
        ip.gauss(path_in, path_out)
    elif filter == 'MedianFilter':
        path_out = out_dir_filter + '/' + filter + '/' + noise + '/' + filename
        ip.median(path_in, path_out)
    elif filter == 'MeanFilter':
        path_out = out_dir_filter + '/' + filter + '/' + noise + '/' + filename
        ip.mean(path_in, path_out)
    elif filter == 'BilateralFilter':
        path_out = out_dir_filter + '/' + filter + '/' + noise + '/' + filename
        ip.bilateral(path_in, path_out)
    elif filter == 'NoFilter':
        path_out = out_dir_filter + '/' + filter + '/' + noise + '/' + filename
        im = cv2.imread(path_in)
        cv2.imwrite(path_out, im)
    else:
        print('Not a valid filter')


def preprocess():
    j = 0
    raw_img_dir = 'raw_input_img'
    img_dir = 'input_img'
    filters = ['NoFilter', 'MedianFilter', 'GaussianFilter', 'MeanFilter', 'BilateralFilter']
    noises = ['NoAddedNoise', 'S&P', 'GaussianNoise', 'Speckle', 'Poisson']
    counter = 0
    total = len(os.listdir(raw_img_dir)) * 375

    while j < len(os.listdir(raw_img_dir)):
           for filename in os.listdir(raw_img_dir):
                path_in = raw_img_dir + '/' + filename
                for noise in noises:
                    add_noise(path_in, out_dir_noise, noise, j)
                    #print(j)
                    counter += 1
                    print('procent done',counter/total * 100)
                j += 1

    for n in noises:
        i = 0
        p = img_dir + '/' + n
        while i < len(os.listdir(p)):
            for filename in os.listdir(p):
                path_in = p + '/' + filename
                for fil in filters:
                    out = out_dir_filter + fil
                    noise = n
                    apply_filter(path_in, out, fil, noise, filename, i)
                    #print(i)
                    counter += 1
                    print('procent done' ,counter / total * 100)
            i += 1


def input_img_getRGB(path):
    spl = path.split('_')
    R = spl[7]
    G = spl[8]
    B = spl[9]
    return R + '_' + G + '_' + B + '_'

def input_img_getHSL(path):
    spl = path.split('_')
    R = float(spl[0])
    G = float(spl[1])
    B = float(spl[2])
    rgb = [R, G, B]
    HLS = colorsys.rgb_to_hls(R, G, B)
    HSL = [HLS[0], HLS[2], HLS[1]]
    return HSL

def getID(path):
    path = path[:-4]
    spl = path.split('_')
    return spl[-1]

def create_img_dict():
    i = 0
    img_list = os.listdir('raw_input_img')
    nb = len(img_list)
    while i < nb:
        coord_dict[i] = []
        i += 1


def colorpicker(image_path, wbook, wsheet, curr):
    image = cv2.imread(image_path)
    start = 2

    eratio = 0.20
    height, width = image.shape[:2]
    nbOfRows = math.ceil(math.sqrt(nbp))
    nbOfCols = math.ceil(nbp / nbOfRows)
    nbOfColors = 1
    eheight = eratio * height
    ewidth = eratio * width
    hsize = (height - eheight * 2) / (nbOfRows + 5)
    wsize = (width - ewidth * 2) / (nbOfCols + 5) / nbOfColors
    wcsize = (width - ewidth * 2) / nbOfColors
    cnbp = 0
    for w in range(3, 3 + nbOfCols):
        for h in range(3, 3 + nbOfRows):
            if cnbp >= nbp:
                break
            x = int(ewidth + wsize * w)
            y = int(eheight + hsize * h)
            R = image[y, x][2]
            G = image[y, x][1]
            B = image[y, x][0]
            HLS = colorsys.rgb_to_hls(R, G, B)
            HSL = [HLS[0], HLS[2], HLS[1]]
            curr_col = str(int(curr) + cnbp)
            wsheet['D' + curr_col] = HSL[0]
            wsheet['E' + curr_col] = HSL[1]
            wsheet['F' + curr_col] = HSL[2]
            cnbp += 1


def pick_color_loop(imgpath, sheet):
    imglist = os.listdir(imgpath)


def create_exc():
    for ws in ws_names:
        current_path = 'images/' + ws
        #print(current_path)
        for wb in wb_names:
            current_folder = current_path + '/' + wb
            #print(current_folder)
            temp_lst = os.listdir(current_folder)
            excelfile = wb + exc
            workb = load_workbook(excelfile)
            wsheet = workb[ws]
            wsheet["A1"] = "H"
            wsheet["B1"] = "S"
            wsheet["C1"] = "L"
            wsheet["D1"] = "H_"
            wsheet["E1"] = "S_"
            wsheet["F1"] = "L_"
            workb.save(excelfile)




current_row = 2

for ws in ws_names:
    current_path = 'images/' + ws
    #print(current_path)
    for wb in wb_names:
        current_folder = current_path + '/' + wb
        #print(current_folder)
        temp_lst = os.listdir(current_folder)
        print(temp_lst)
        excelfile = wb + exc
        workb = load_workbook(excelfile)
        wsheet = workb[ws]
        current_row = 2
        while current_row < (len(temp_lst) + 2):
            rownb = str(current_row)
            current_pic_path = temp_lst[current_row-2]
            cp = 0
            while cp < nbp:
                wsheet['A' + str(current_row+cp)] = input_img_getHSL(current_pic_path)[0]
                wsheet['B' + str(current_row+cp)] = input_img_getHSL(current_pic_path)[1]
                wsheet['C' + str(current_row+cp)] = input_img_getHSL(current_pic_path)[2]
                cp += 1
            print(current_folder + current_pic_path)
            colorpicker(current_folder + '/' + current_pic_path, workb, wsheet, rownb)
            current_row += 1

        workb.save(excelfile)



#create_img_dict()
#preprocess()


'''
---- IMAGEPICKER ----

# Create empty list for coordinate arrays to be appended to
coords = []

# Function to be called when mouse is clicked
def save_coords(event, id):
    click_loc = (event.x, event.y)
    print("you clicked on", click_loc)
    coord_dict[id].append(click_loc)


# Function to load the next image into the Label
def next_img():
    n = Image.open(next(imgs))
    n = n.resize((480, 720), Image.ANTIALIAS)
    img_label.img = ImageTk.PhotoImage(n)
    img_label.config(image=img_label.img)

root = tk.Tk()
root.pack_slaves()

# Choose multiple images
img_dir = filedialog.askdirectory(parent=root, initialdir="test2-colors", title='Choose folder')
os.chdir(img_dir)
imgs = iter(os.listdir(img_dir))


img_label = tk.Label(root)
img_label.pack(side = "left")
img_label.pack(expand=0)
img_label.bind("<Button-1>",save_coords)

#btn = tk.Button(root, text='Next image', command=next_img)
#btn.pack(side='right')

next_img() # load first image

root.mainloop()

print(coords)
'''