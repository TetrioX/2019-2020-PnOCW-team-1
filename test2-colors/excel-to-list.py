# Dependency: openpyxl

from openpyxl import load_workbook, cell as xlcell

# excel file name
fileName = "resultAll.xlsx"

# filters (None means doesnt filter)
group = None # group
brigthness = None # brigthness
pictureDevice = None # pictureDevice
ratioScreenPicture = None # ratioScreenPicture
distance = None # distance
nbOfColros = None # nbOfColros
incidence = lambda x : True # incidence
reflection = None # reflection
screenDevice = None # screenDevice
imageID = None # imageID
origHexCol = "#00ffff"
neigborColsf = lambda x : x == "#00ff00"

def key(row):
    return row[-1]
    ind = row[0]
    if ind == "0":
        return "geen licht"
    if ind == "1":
        return "artificieel licht"
    if ind == "2":
        return "natuurlijk licht"

global ws
neigbors = []
neigborCols = set()
result = []
keys = []
print("==> reading resultAll.xlsx (this can take several minutes)...", flush=True)
wb = load_workbook(filename = './'+"resultAll.xlsx", read_only=True, data_only=True)
ws = wb.active
def checkFilters(values, filters, row):
    global ws
    prevCol = row[10].value
    for i in range(len(filters)):
        val = row[filters[i][0]].value
        if val is not None:
            values[i] = val
        if callable(filters[i][1]):
            if not filters[i][1](values[i]):
                return False
        else:
            if values[i] != filters[i][1] and filters[i][1] is not None:
                return False
    return True


# setting up filters variables
print("==> setting up filters", flush=True)
image = 2
allFilters = [group, brigthness, pictureDevice, ratioScreenPicture, distance, \
nbOfColros, incidence, reflection, screenDevice, imageID, origHexCol]
filters = []
for i in range(len(allFilters)):
    if allFilters[i] is not None or i == 10:
        filters.append((i, allFilters[i]))
curValues = [None]*len(filters)

print("==> filtering values", flush=True)
prevHash = None
hash = None
for row in ws.iter_rows(min_row=2, max_col=12):
    if row[11].value is None:
        break
    if callable(neigborColsf):
        if row[9].value is not None:
            hash = row[9].value
        if prevHash != hash:
            neigborCols.discard(None)
            for col in neigborCols:
                if neigborColsf(col):
                    for row in neigbors:
                        if checkFilters(curValues, filters, row):
                            result.append(row[11].value)
                            keys.append(key(curValues))
                    continue
            neigbors = []
            neigborCols = set()
        prevHash = hash
        neigbors.append(row)
        neigborCols.add(row[10].value)
    else:
        if checkFilters(curValues, filters, row):
            result.append(row[11].value)
            keys.append(key(curValues))



with open('result.txt', 'w') as filehandle:
    for listitem in result:
        filehandle.write('%s\n' % listitem)

with open('keys.txt', 'w') as filehandle:
    for listitem in keys:
        filehandle.write('%s\n' % listitem)
