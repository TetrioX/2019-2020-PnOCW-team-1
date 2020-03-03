from openpyxl import Workbook, load_workbook, cell as xlcell, worksheet


def hexToRGB(h): # tuple (R,G,B)
    h = h.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))




global ws
condition = [None, None]
result = []
keys = []
wb = load_workbook(filename = './result.xlsx')
print("Hoppaaaaaa")
ws = wb.active

image = 2


def within_range(bounds: tuple, cell: xlcell) -> bool:
    column_start, row_start, column_end, row_end = bounds
    row = cell.row
    if row >= row_start and row <= row_end:
        column = cell.column
        if column >= column_start and column <= column_end:
            return True
    return False

def get_value_merged(sheet: worksheet, cell: xlcell) -> any:
    for merged in sheet.merged_cells:
        if within_range(merged.bounds, cell):
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value

def select_brightness(value):
    condition[0] = "B"
    condition[1] = value

def select_distance(value):
    condition[0] = "E"
    condition[1] = value

def select_nb_colors(value):
    condition[0] = "F"
    condition[1] = value

def select_reflection(value):
    condition[0] = "H"
    condition[1] = value

def select_incidence(value):
    condition[0] = "G"
    condition[1] = value


def read_values():
    i = 2
    cond = condition[0]
    val = condition[1]
    print("cond=", cond)
    print('val=', val)
    while ws["L" + str(i)].value is not None:
        if get_value_merged(ws, ws[cond + str(i)]) == str(val):
            result.append(ws["L" + str(i)].value)
            keys.append(get_value_merged(ws, ws["K" + str(i)]))
        i += 1


select_brightness(50)
print("condition = ", condition)
read_values()
print(result)



