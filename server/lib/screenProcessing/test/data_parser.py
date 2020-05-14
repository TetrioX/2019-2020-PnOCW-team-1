import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Series, Reference
import json
import time


cs = 'caseNB'
scr = 'screensFound'
sq = 'squaresFound'
tme = 'time'
n = 12

#TOT EN MET CASE 44
total_screens = 176
total_squares = 1298
jpath = './results/results0.json'
expath = './results/results0.xlsx'

def get_amount_of_sheets(wb):
    count = 0
    for sheet in wb:
        count += 1
    return count

def initExcel(nb, excel_path):
    wb = Workbook()
    for i in range(nb):
        sh_name = 'Tresholds' + str(i)
        wb.create_sheet(sh_name)
        ws = wb[sh_name]
        ws['A1'] = 'Case Number'
        ws['B1'] = 'Screens found'
        ws['C1'] = 'Squares found'
        ws['D1'] = 'Time elapsed'
        ws['F1'] = 'Totaal schermen'
        ws['H1'] = 'Totaal vierkanten'
        ws['K1'] = 'Percentage schermen'
        ws['N1'] = 'Percentage vierkanten'
        ws['Q1'] = 'Gemiddelde tijd'

    wb.save(excel_path)

def writeExcel(nb, excel_path, json_path):
    with open(json_path, encoding="utf8") as f:
        json_data = json.load(f)
    wb = load_workbook(excel_path)
    for i in range(nb):
        j = 1
        sh_name = 'Tresholds' + str(i)
        ws = wb[sh_name]
        val = json_data[str(i)]
        mx = ws.max_row
        for elem in val:
            ws['A' + str(mx+j)] = elem[cs]
            ws['B' + str(mx + j)] = elem[scr]
            ws['C' + str(mx + j)] = elem[sq]
            ws['D' + str(mx + j)] = elem[tme]
            j += 1
        mx2 = ws.max_row
        totalB = 0
        totalC = 0
        totalD = []
        for row in ws.iter_rows(min_row=2, max_row=mx2,min_col=2,max_col=2):
            for cell in row:
                totalB += cell.value
        for row in ws.iter_rows(min_row=2, max_row=mx2,min_col=3,max_col=3):
            for cell in row:
                totalC += cell.value
        for row in ws.iter_rows(min_row=2, max_row=mx2,min_col=4,max_col=4):
            for cell in row:
                totalD.append(cell.value)
        totalD.sort()
        ws['F2'] = totalB
        ws['H2'] = totalC
        ws['K2'] = (100*totalB)/total_screens
        ws['N2'] = (100*totalC)/total_squares
        ws['Q2'] = totalD[len(totalD)//2]

    wb.save(excel_path)

def createChart(excel_path):
    rows = [('Used Tresholds', '% Screens', '% Squares', 'Time elapsed')]
    wb = load_workbook(excel_path)
    ws = wb.active
    print(ws)
    for sheet in wb:
        print(sheet)
        curr_ws = sheet
        if ws != curr_ws:
            result = (curr_ws.title, curr_ws['K2'].value, curr_ws['N2'].value, curr_ws['Q2'].value)
            rows.append(result)
    ws = wb.active
    for row in rows:
        ws.append(row)
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Resultaten testcases"
    chart1.y_axis.title = 'Percentage results'
    chart1.x_axis.title = 'Tresholds used'
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4

    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Mediaan tijd testcases"
    chart2.y_axis.title = 'Mediaan tijd'
    chart2.x_axis.title = 'Tresholds used'
    data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row, max_col=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.shape = 4

    ws.add_chart(chart1, "A20")
    ws.add_chart(chart2, "J20")
    wb.save(excel_path)

def main():
    initExcel(n,expath)
    writeExcel(n,expath,jpath)
    createChart(expath)


main()